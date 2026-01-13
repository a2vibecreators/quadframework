#!/usr/bin/env python3
"""
QUAD Excel Template Generator
=============================

Creates an interactive Excel template with:
- Dropdowns linked to data
- Validation rules
- Conditional formatting
- Professional styling

Copyright (c) 2026 Gopi Suman Addanke. All Rights Reserved.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


def create_quad_template(filename: str = "quad-org-template.xlsx"):
    """Create the QUAD organization Excel template"""

    wb = Workbook()

    # ─────────────────────────────────────────────────────────
    # Styles
    # ─────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Light yellow
    locked_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # Gray
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ─────────────────────────────────────────────────────────
    # Tab 1: Overview
    # ─────────────────────────────────────────────────────────
    ws_overview = wb.active
    ws_overview.title = "Overview"

    # Title
    ws_overview.merge_cells('A1:C1')
    ws_overview['A1'] = "QUAD Organization Setup"
    ws_overview['A1'].font = Font(bold=True, size=16, color="2B579A")

    ws_overview['A2'] = "Fill in yellow cells. Dropdowns provided where applicable."
    ws_overview['A2'].font = Font(italic=True, color="666666")

    # Headers
    overview_fields = [
        ("Organization Name", "Your company or team name", "Acme Corp"),
        ("Organization Code", "Short code (3-6 chars, used in URLs)", "ACME"),
        ("Timezone", "Select timezone", "America/New_York"),
        ("Industry", "Select industry", "Technology"),
        ("Contact Email", "Primary contact email", "admin@acme.com"),
    ]

    row = 4
    ws_overview['A4'] = "Field"
    ws_overview['B4'] = "Value"
    ws_overview['C4'] = "Notes"
    for col in ['A', 'B', 'C']:
        ws_overview[f'{col}4'].font = header_font
        ws_overview[f'{col}4'].fill = header_fill
        ws_overview[f'{col}4'].border = thin_border

    for i, (field, note, example) in enumerate(overview_fields, start=5):
        ws_overview[f'A{i}'] = field
        ws_overview[f'A{i}'].font = Font(bold=True)
        ws_overview[f'A{i}'].border = thin_border

        ws_overview[f'B{i}'].fill = input_fill
        ws_overview[f'B{i}'].border = thin_border
        ws_overview[f'B{i}'].value = example

        ws_overview[f'C{i}'] = note
        ws_overview[f'C{i}'].font = Font(italic=True, color="666666")
        ws_overview[f'C{i}'].border = thin_border

    # Timezone dropdown
    tz_validation = DataValidation(
        type="list",
        formula1='"America/New_York,America/Chicago,America/Denver,America/Los_Angeles,America/Phoenix,UTC,Europe/London,Asia/Tokyo,Asia/Singapore"',
        allow_blank=True
    )
    tz_validation.error = "Please select a valid timezone"
    tz_validation.errorTitle = "Invalid Timezone"
    ws_overview.add_data_validation(tz_validation)
    tz_validation.add('B6')

    # Industry dropdown
    industry_validation = DataValidation(
        type="list",
        formula1='"Technology,Healthcare,Finance,Retail,Manufacturing,Education,Government,Non-Profit,Other"',
        allow_blank=True
    )
    ws_overview.add_data_validation(industry_validation)
    industry_validation.add('B7')

    # Column widths
    ws_overview.column_dimensions['A'].width = 25
    ws_overview.column_dimensions['B'].width = 30
    ws_overview.column_dimensions['C'].width = 40

    # ─────────────────────────────────────────────────────────
    # Tab 2: Resources (Team)
    # ─────────────────────────────────────────────────────────
    ws_resources = wb.create_sheet("Resources")

    # Title
    ws_resources.merge_cells('A1:G1')
    ws_resources['A1'] = "Team Resources"
    ws_resources['A1'].font = Font(bold=True, size=16, color="2B579A")

    ws_resources['A2'] = "Add team members. Allocation % must total ≤100 per person across projects."
    ws_resources['A2'].font = Font(italic=True, color="666666")

    # Headers
    resource_headers = ["Name", "Email", "Role", "Skills", "Allocation %", "Available Hrs", "Notes"]
    for col, header in enumerate(resource_headers, start=1):
        cell = ws_resources.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Role dropdown
    role_validation = DataValidation(
        type="list",
        formula1='"DIRECTOR,MANAGER,TECH_LEAD,SENIOR_DEVELOPER,DEVELOPER,JUNIOR_DEVELOPER,QA_ENGINEER,DESIGNER,DEVOPS,PRODUCT_MANAGER,SCRUM_MASTER"',
        allow_blank=False
    )
    role_validation.error = "Please select a valid role"
    role_validation.errorTitle = "Invalid Role"
    role_validation.prompt = "Select team member role"
    role_validation.promptTitle = "Role"
    ws_resources.add_data_validation(role_validation)

    # Allocation % validation (0-100)
    allocation_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="100",
        allow_blank=True
    )
    allocation_validation.error = "Allocation must be between 0 and 100"
    allocation_validation.errorTitle = "Invalid Allocation"
    allocation_validation.prompt = "Enter allocation percentage (0-100)"
    allocation_validation.promptTitle = "Allocation %"
    ws_resources.add_data_validation(allocation_validation)

    # Sample data rows
    sample_resources = [
        ("Suman Addanki", "suman@acme.com", "DIRECTOR", "Java, Python, Architecture", 100, 0, "Founder"),
        ("Dev One", "dev1@acme.com", "SENIOR_DEVELOPER", "React, Node.js, TypeScript", 80, 8, "Frontend lead"),
        ("Dev Two", "dev2@acme.com", "DEVELOPER", "Java, Spring Boot, PostgreSQL", 60, 16, "Backend"),
        ("", "", "", "", "", "", ""),
        ("", "", "", "", "", "", ""),
    ]

    for row_num, resource in enumerate(sample_resources, start=5):
        for col_num, value in enumerate(resource, start=1):
            cell = ws_resources.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if col_num in [1, 2, 3, 4, 7]:  # Text columns
                cell.fill = input_fill
            elif col_num == 5:  # Allocation
                cell.fill = input_fill
                cell.alignment = Alignment(horizontal='center')
            elif col_num == 6:  # Available Hrs (calculated)
                cell.fill = locked_fill
                cell.alignment = Alignment(horizontal='center')
                if value != "":
                    # Formula: 40 * (100 - allocation) / 100
                    cell.value = f"=40*(100-E{row_num})/100" if row_num <= 7 else ""

    # Apply validations to data range
    role_validation.add('C5:C50')
    allocation_validation.add('E5:E50')

    # Conditional formatting for allocation (red if > 100 total)
    ws_resources.conditional_formatting.add(
        'E5:E50',
        FormulaRule(
            formula=['E5>100'],
            fill=error_fill
        )
    )

    # Column widths
    ws_resources.column_dimensions['A'].width = 20
    ws_resources.column_dimensions['B'].width = 25
    ws_resources.column_dimensions['C'].width = 18
    ws_resources.column_dimensions['D'].width = 30
    ws_resources.column_dimensions['E'].width = 14
    ws_resources.column_dimensions['F'].width = 14
    ws_resources.column_dimensions['G'].width = 25

    # ─────────────────────────────────────────────────────────
    # Tab 3: Project 1
    # ─────────────────────────────────────────────────────────
    ws_project1 = wb.create_sheet("Project 1")
    setup_project_sheet(ws_project1, "SUMA Platform", "SUMA AI Platform Development",
                        header_font, header_fill, input_fill, thin_border, wb)

    # ─────────────────────────────────────────────────────────
    # Tab 4: Project 2
    # ─────────────────────────────────────────────────────────
    ws_project2 = wb.create_sheet("Project 2")
    setup_project_sheet(ws_project2, "NutriNine", "Nutrition tracking application",
                        header_font, header_fill, input_fill, thin_border, wb)

    # ─────────────────────────────────────────────────────────
    # Tab 5: Features (for quad-plan)
    # ─────────────────────────────────────────────────────────
    ws_features = wb.create_sheet("Features")

    ws_features.merge_cells('A1:G1')
    ws_features['A1'] = "Project Features"
    ws_features['A1'].font = Font(bold=True, size=16, color="2B579A")

    ws_features['A2'] = "Define features for PGCE prioritization. Dependencies determine build order."
    ws_features['A2'].font = Font(italic=True, color="666666")

    feature_headers = ["Feature ID", "Feature Name", "Description", "Business Value (1-10)", "Complexity (1-10)", "Depends On", "Project"]
    for col, header in enumerate(feature_headers, start=1):
        cell = ws_features.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Business Value validation (1-10)
    bv_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="10",
        allow_blank=False
    )
    bv_validation.error = "Business Value must be 1-10"
    bv_validation.prompt = "1=Low, 10=Critical"
    bv_validation.promptTitle = "Business Value"
    ws_features.add_data_validation(bv_validation)

    # Complexity validation (1-10)
    cx_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="10",
        allow_blank=False
    )
    cx_validation.error = "Complexity must be 1-10"
    cx_validation.prompt = "1=Simple, 10=Very Complex"
    cx_validation.promptTitle = "Complexity"
    ws_features.add_data_validation(cx_validation)

    # Sample features for Bank Demo
    sample_features = [
        ("F001", "Account Management", "Create, view, update bank accounts", 9, 4, "", "Bank Demo"),
        ("F002", "Transaction Processing", "Deposits, withdrawals, transfers", 10, 7, "F001", "Bank Demo"),
        ("F003", "Balance Inquiry", "Check balance and transaction history", 8, 3, "F001,F002", "Bank Demo"),
        ("F004", "User Authentication", "Login, logout, session management", 10, 5, "", "Bank Demo"),
        ("", "", "", "", "", "", ""),
        ("", "", "", "", "", "", ""),
    ]

    for row_num, feature in enumerate(sample_features, start=5):
        for col_num, value in enumerate(feature, start=1):
            cell = ws_features.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.fill = input_fill
            if col_num in [4, 5]:  # Numeric columns
                cell.alignment = Alignment(horizontal='center')

    bv_validation.add('D5:D50')
    cx_validation.add('E5:E50')

    # Column widths
    ws_features.column_dimensions['A'].width = 12
    ws_features.column_dimensions['B'].width = 25
    ws_features.column_dimensions['C'].width = 40
    ws_features.column_dimensions['D'].width = 18
    ws_features.column_dimensions['E'].width = 18
    ws_features.column_dimensions['F'].width = 15
    ws_features.column_dimensions['G'].width = 15

    # ─────────────────────────────────────────────────────────
    # Tab 6: Integrations (external services)
    # ─────────────────────────────────────────────────────────
    ws_integrations = wb.create_sheet("Integrations")

    ws_integrations.merge_cells('A1:F1')
    ws_integrations['A1'] = "External Integrations"
    ws_integrations['A1'].font = Font(bold=True, size=16, color="2B579A")

    ws_integrations['A2'] = "Define external APIs and services the project connects to."
    ws_integrations['A2'].font = Font(italic=True, color="666666")

    int_headers = ["Service Name", "Type", "Base URL", "Auth Method", "Project", "Notes"]
    for col, header in enumerate(int_headers, start=1):
        cell = ws_integrations.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Type dropdown
    int_type_validation = DataValidation(
        type="list",
        formula1='"REST API,GraphQL,gRPC,WebSocket,Database,Message Queue,File Storage,Email/SMS,Payment Gateway,Other"',
        allow_blank=False
    )
    ws_integrations.add_data_validation(int_type_validation)

    # Auth dropdown
    auth_validation = DataValidation(
        type="list",
        formula1='"None,API Key,OAuth 2.0,JWT,Basic Auth,mTLS,IAM Role"',
        allow_blank=False
    )
    ws_integrations.add_data_validation(auth_validation)

    # Sample integrations
    sample_integrations = [
        ("PostgreSQL", "Database", "localhost:5432", "Basic Auth", "Bank Demo", "Primary database"),
        ("Email Service", "Email/SMS", "smtp.gmail.com", "OAuth 2.0", "Bank Demo", "Notifications"),
        ("", "", "", "", "", ""),
    ]

    for row_num, integration in enumerate(sample_integrations, start=5):
        for col_num, value in enumerate(integration, start=1):
            cell = ws_integrations.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.fill = input_fill

    int_type_validation.add('B5:B30')
    auth_validation.add('D5:D30')

    # Column widths
    ws_integrations.column_dimensions['A'].width = 20
    ws_integrations.column_dimensions['B'].width = 15
    ws_integrations.column_dimensions['C'].width = 30
    ws_integrations.column_dimensions['D'].width = 15
    ws_integrations.column_dimensions['E'].width = 15
    ws_integrations.column_dimensions['F'].width = 30

    # ─────────────────────────────────────────────────────────
    # Tab 7: Security Requirements
    # ─────────────────────────────────────────────────────────
    ws_security = wb.create_sheet("Security")

    ws_security.merge_cells('A1:D1')
    ws_security['A1'] = "Security & Compliance"
    ws_security['A1'].font = Font(bold=True, size=16, color="2B579A")

    ws_security['A2'] = "Define security requirements and compliance standards."
    ws_security['A2'].font = Font(italic=True, color="666666")

    # Security settings
    security_fields = [
        ("Authentication Method", "JWT", "JWT,OAuth 2.0,Session,API Key,SAML,None"),
        ("Authorization Model", "RBAC", "RBAC,ABAC,ACL,None"),
        ("Data Encryption (at rest)", "AES-256", "AES-256,AES-128,None"),
        ("Data Encryption (in transit)", "TLS 1.3", "TLS 1.3,TLS 1.2,None"),
        ("Password Policy", "Strong", "Strong,Medium,Basic"),
        ("Session Timeout (minutes)", "30", None),
        ("MFA Required", "No", "Yes,No,Optional"),
    ]

    ws_security['A4'] = "Setting"
    ws_security['B4'] = "Value"
    ws_security['C4'] = "Notes"
    for col in ['A', 'B', 'C']:
        ws_security[f'{col}4'].font = header_font
        ws_security[f'{col}4'].fill = header_fill
        ws_security[f'{col}4'].border = thin_border

    for i, (setting, default, options) in enumerate(security_fields, start=5):
        ws_security[f'A{i}'] = setting
        ws_security[f'A{i}'].font = Font(bold=True)
        ws_security[f'A{i}'].border = thin_border

        ws_security[f'B{i}'] = default
        ws_security[f'B{i}'].fill = input_fill
        ws_security[f'B{i}'].border = thin_border

        ws_security[f'C{i}'].border = thin_border

        if options:
            dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=False)
            ws_security.add_data_validation(dv)
            dv.add(f'B{i}')

    # Compliance Section
    ws_security['A13'] = "Compliance Requirements"
    ws_security['A13'].font = Font(bold=True, size=12)

    compliance_options = [
        ("SOC 2", "No"),
        ("HIPAA", "No"),
        ("PCI-DSS", "No"),
        ("GDPR", "No"),
        ("ISO 27001", "No"),
    ]

    yes_no_sec = DataValidation(type="list", formula1='"Yes,No,In Progress"', allow_blank=False)
    ws_security.add_data_validation(yes_no_sec)

    for i, (compliance, default) in enumerate(compliance_options, start=14):
        ws_security[f'A{i}'] = compliance
        ws_security[f'A{i}'].border = thin_border
        ws_security[f'B{i}'] = default
        ws_security[f'B{i}'].fill = input_fill
        ws_security[f'B{i}'].border = thin_border
        yes_no_sec.add(f'B{i}')

    # Column widths
    ws_security.column_dimensions['A'].width = 28
    ws_security.column_dimensions['B'].width = 20
    ws_security.column_dimensions['C'].width = 35

    # ─────────────────────────────────────────────────────────
    # Tab 8: Secrets (environment variables)
    # ─────────────────────────────────────────────────────────
    ws_secrets = wb.create_sheet("Secrets")

    ws_secrets.merge_cells('A1:E1')
    ws_secrets['A1'] = "Secrets & Environment Variables"
    ws_secrets['A1'].font = Font(bold=True, size=16, color="2B579A")

    ws_secrets['A2'] = "Define secrets needed for deployment. DO NOT put actual values here!"
    ws_secrets['A2'].font = Font(italic=True, color="CC0000")

    secret_headers = ["Secret Name", "Environment Variable", "Source", "Environment", "Description"]
    for col, header in enumerate(secret_headers, start=1):
        cell = ws_secrets.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Source dropdown
    source_validation = DataValidation(
        type="list",
        formula1='"GCP Secret Manager,AWS Secrets Manager,Azure Key Vault,HashiCorp Vault,Environment File,Manual"',
        allow_blank=False
    )
    ws_secrets.add_data_validation(source_validation)

    # Environment dropdown
    secret_env_validation = DataValidation(
        type="list",
        formula1='"all,dev,staging,prod"',
        allow_blank=False
    )
    ws_secrets.add_data_validation(secret_env_validation)

    # Sample secrets
    sample_secrets = [
        ("db-password", "DB_PASSWORD", "GCP Secret Manager", "all", "Database password"),
        ("jwt-secret", "JWT_SECRET", "GCP Secret Manager", "all", "JWT signing key"),
        ("api-key", "EXTERNAL_API_KEY", "GCP Secret Manager", "prod", "Third-party API key"),
        ("", "", "", "", ""),
    ]

    for row_num, secret in enumerate(sample_secrets, start=5):
        for col_num, value in enumerate(secret, start=1):
            cell = ws_secrets.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.fill = input_fill

    source_validation.add('C5:C30')
    secret_env_validation.add('D5:D30')

    # Column widths
    ws_secrets.column_dimensions['A'].width = 20
    ws_secrets.column_dimensions['B'].width = 25
    ws_secrets.column_dimensions['C'].width = 22
    ws_secrets.column_dimensions['D'].width = 12
    ws_secrets.column_dimensions['E'].width = 35

    # ─────────────────────────────────────────────────────────
    # Tab 9: Deployment
    # ─────────────────────────────────────────────────────────
    ws_deploy = wb.create_sheet("Deployment")

    ws_deploy.merge_cells('A1:F1')
    ws_deploy['A1'] = "Deployment Configuration"
    ws_deploy['A1'].font = Font(bold=True, size=16, color="2B579A")

    ws_deploy['A2'] = "Configure deployment targets for each project and environment."
    ws_deploy['A2'].font = Font(italic=True, color="666666")

    deploy_headers = ["Project", "Environment", "GCP Project ID", "Region", "Git Repo", "Service Name"]
    for col, header in enumerate(deploy_headers, start=1):
        cell = ws_deploy.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Environment dropdown
    env_validation = DataValidation(
        type="list",
        formula1='"dev,staging,prod"',
        allow_blank=False
    )
    ws_deploy.add_data_validation(env_validation)

    # Region dropdown
    region_validation = DataValidation(
        type="list",
        formula1='"us-central1,us-east1,us-west1,europe-west1,asia-east1,asia-southeast1"',
        allow_blank=False
    )
    ws_deploy.add_data_validation(region_validation)

    # Sample deployment data
    deploy_data = [
        ("SUMA Platform", "dev", "suma-dev", "us-central1", "github.com/a2vibes/quad-suma-api", "suma-api"),
        ("SUMA Platform", "prod", "suma-prod", "us-central1", "github.com/a2vibes/quad-suma-api", "suma-api"),
        ("NutriNine", "dev", "nutrinne-dev", "us-central1", "github.com/a2vibes/nutrinne-api", "nutrinne-api"),
        ("NutriNine", "prod", "nutrinne-prod", "us-central1", "github.com/a2vibes/nutrinne-api", "nutrinne-api"),
    ]

    for row_num, data in enumerate(deploy_data, start=5):
        for col_num, value in enumerate(data, start=1):
            cell = ws_deploy.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.fill = input_fill

    env_validation.add('B5:B20')
    region_validation.add('D5:D20')

    # Column widths
    ws_deploy.column_dimensions['A'].width = 18
    ws_deploy.column_dimensions['B'].width = 12
    ws_deploy.column_dimensions['C'].width = 18
    ws_deploy.column_dimensions['D'].width = 15
    ws_deploy.column_dimensions['E'].width = 40
    ws_deploy.column_dimensions['F'].width = 18

    # ─────────────────────────────────────────────────────────
    # Tab 10: Lookups (Hidden data for dropdowns)
    # ─────────────────────────────────────────────────────────
    ws_lookups = wb.create_sheet("_Lookups")

    # This sheet contains reference data for dropdowns
    ws_lookups['A1'] = "Project Types"
    project_types = ["Web Application", "Mobile App", "API/Backend", "Desktop App", "Library/Package", "Microservice"]
    for i, pt in enumerate(project_types, start=2):
        ws_lookups[f'A{i}'] = pt

    ws_lookups['B1'] = "Frontend Tech"
    frontend_tech = ["React.js", "Next.js", "Vue.js", "Angular", "Swift (iOS)", "Kotlin (Android)", "Flutter", "None"]
    for i, ft in enumerate(frontend_tech, start=2):
        ws_lookups[f'B{i}'] = ft

    ws_lookups['C1'] = "Backend Tech"
    backend_tech = ["Node.js", "Python/FastAPI", "Java/Spring Boot", "Go", ".NET Core", "Ruby on Rails", "None"]
    for i, bt in enumerate(backend_tech, start=2):
        ws_lookups[f'C{i}'] = bt

    ws_lookups['D1'] = "Database"
    databases = ["PostgreSQL", "MySQL", "MongoDB", "SQLite", "Redis", "DynamoDB", "None"]
    for i, db in enumerate(databases, start=2):
        ws_lookups[f'D{i}'] = db

    ws_lookups['E1'] = "Deliverables"
    deliverables = ["Web Application", "API Server", "Mobile App (iOS)", "Mobile App (Android)",
                    "JAR File", "Docker Image", "Documentation", "SDK/Library"]
    for i, d in enumerate(deliverables, start=2):
        ws_lookups[f'E{i}'] = d

    # Define named ranges for lookups
    wb.defined_names["ProjectTypes"] = DefinedName(
        name="ProjectTypes",
        attr_text=f"'_Lookups'!$A$2:$A${len(project_types)+1}"
    )
    wb.defined_names["FrontendTech"] = DefinedName(
        name="FrontendTech",
        attr_text=f"'_Lookups'!$B$2:$B${len(frontend_tech)+1}"
    )
    wb.defined_names["BackendTech"] = DefinedName(
        name="BackendTech",
        attr_text=f"'_Lookups'!$C$2:$C${len(backend_tech)+1}"
    )
    wb.defined_names["Databases"] = DefinedName(
        name="Databases",
        attr_text=f"'_Lookups'!$D$2:$D${len(databases)+1}"
    )

    # Create named range for Resources (team member names)
    wb.defined_names["TeamMembers"] = DefinedName(
        name="TeamMembers",
        attr_text="'Resources'!$A$5:$A$50"
    )

    # Hide the lookups sheet
    ws_lookups.sheet_state = 'hidden'

    # ─────────────────────────────────────────────────────────
    # Save
    # ─────────────────────────────────────────────────────────
    wb.save(filename)
    print(f"✓ Created template: {filename}")
    print("")
    print("  Tabs:")
    print("  1. Overview      - Organization details")
    print("  2. Resources     - Team members & skills")
    print("  3. Project 1     - First project config")
    print("  4. Project 2     - Second project config")
    print("  5. Features      - Feature definitions (for quad-plan)")
    print("  6. Integrations  - External services")
    print("  7. Security      - Security & compliance")
    print("  8. Secrets       - Environment variables")
    print("  9. Deployment    - Deploy configurations")
    print("")
    print("  Features:")
    print("  • Dropdowns for all selections")
    print("  • PGCE inputs: Business Value, Complexity, Dependencies")
    print("  • Yellow cells = editable, Gray cells = calculated")


def setup_project_sheet(ws, project_name, description, header_font, header_fill, input_fill, thin_border, wb):
    """Setup a project sheet with validations"""

    ws.merge_cells('A1:C1')
    ws['A1'] = f"Project: {project_name}"
    ws['A1'].font = Font(bold=True, size=16, color="2B579A")

    # Basic Info Section
    ws['A3'] = "Basic Information"
    ws['A3'].font = Font(bold=True, size=12)

    fields = [
        ("Project Name", project_name, "B4"),
        ("Description", description, "B5"),
        ("Project Type", "", "B6"),
        ("Start Date", "", "B7"),
        ("Deadline", "", "B8"),
        ("Project Owner", "", "B9"),
    ]

    for i, (label, value, cell_ref) in enumerate(fields):
        row = 4 + i
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = thin_border
        ws[cell_ref] = value
        ws[cell_ref].fill = input_fill
        ws[cell_ref].border = thin_border

    # Tech Stack Section
    ws['A11'] = "Tech Stack"
    ws['A11'].font = Font(bold=True, size=12)

    tech_fields = [
        ("Frontend", "", "B12"),
        ("Backend", "", "B13"),
        ("Database", "", "B14"),
    ]

    for i, (label, value, cell_ref) in enumerate(tech_fields):
        row = 12 + i
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = thin_border
        ws[cell_ref] = value
        ws[cell_ref].fill = input_fill
        ws[cell_ref].border = thin_border

    # Deliverables Section
    ws['A16'] = "Deliverables (select all that apply)"
    ws['A16'].font = Font(bold=True, size=12)

    deliverable_options = [
        ("Web Application", "B17"),
        ("API Server", "B18"),
        ("Mobile App", "B19"),
        ("JAR File", "B20"),
        ("Docker Image", "B21"),
        ("Documentation", "B22"),
    ]

    # Yes/No dropdown for deliverables
    yes_no_validation = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True
    )
    ws.add_data_validation(yes_no_validation)

    for label, cell_ref in deliverable_options:
        row = int(cell_ref[1:])
        ws[f'A{row}'] = label
        ws[f'A{row}'].border = thin_border
        ws[cell_ref] = "No"
        ws[cell_ref].fill = input_fill
        ws[cell_ref].border = thin_border
        yes_no_validation.add(cell_ref)

    # Project Type dropdown
    type_validation = DataValidation(
        type="list",
        formula1="ProjectTypes",
        allow_blank=False
    )
    type_validation.prompt = "Select project type"
    type_validation.promptTitle = "Project Type"
    ws.add_data_validation(type_validation)
    type_validation.add('B6')

    # Frontend dropdown
    fe_validation = DataValidation(
        type="list",
        formula1="FrontendTech",
        allow_blank=True
    )
    ws.add_data_validation(fe_validation)
    fe_validation.add('B12')

    # Backend dropdown
    be_validation = DataValidation(
        type="list",
        formula1="BackendTech",
        allow_blank=True
    )
    ws.add_data_validation(be_validation)
    be_validation.add('B13')

    # Database dropdown
    db_validation = DataValidation(
        type="list",
        formula1="Databases",
        allow_blank=True
    )
    ws.add_data_validation(db_validation)
    db_validation.add('B14')

    # Project Owner dropdown (linked to Resources tab)
    owner_validation = DataValidation(
        type="list",
        formula1="TeamMembers",
        allow_blank=False
    )
    owner_validation.prompt = "Select from team members"
    owner_validation.promptTitle = "Project Owner"
    ws.add_data_validation(owner_validation)
    owner_validation.add('B9')

    # Date validation
    date_validation = DataValidation(
        type="date",
        allow_blank=True
    )
    date_validation.prompt = "Enter date (YYYY-MM-DD)"
    ws.add_data_validation(date_validation)
    date_validation.add('B7')
    date_validation.add('B8')

    # Team Assignment Section
    ws['A24'] = "Team Assignment"
    ws['A24'].font = Font(bold=True, size=12)

    team_headers = ["Team Member", "Role on Project", "Allocation %", "Start Date", "End Date"]
    for col, header in enumerate(team_headers, start=1):
        cell = ws.cell(row=25, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Team member dropdown for assignment
    member_validation = DataValidation(
        type="list",
        formula1="TeamMembers",
        allow_blank=True
    )
    ws.add_data_validation(member_validation)
    member_validation.add('A26:A35')

    # Allocation validation for project
    alloc_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="100",
        allow_blank=True
    )
    alloc_validation.error = "Allocation must be 0-100"
    ws.add_data_validation(alloc_validation)
    alloc_validation.add('C26:C35')

    # Add empty rows for team assignment
    for row in range(26, 31):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = input_fill

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15


if __name__ == "__main__":
    import sys
    filename = sys.argv[1] if len(sys.argv) > 1 else "quad-org-template.xlsx"
    create_quad_template(filename)
