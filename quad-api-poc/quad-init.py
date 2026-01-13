#!/usr/bin/env python3
"""
QUAD Interactive Project Initialization
========================================

Initialize projects from Excel or interactively.

Usage:
  quad-init                      # Interactive mode
  quad-init @org-setup.xlsx      # From Excel
  quad-init --resume bank-demo   # Resume draft

Copyright (c) 2026 Gopi Suman Addanke. All Rights Reserved.
"""

import sys
import os
import json
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any

# Add parent to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import psycopg
    from psycopg.rows import dict_row
    HAS_PSYCOPG = True
except ImportError:
    HAS_PSYCOPG = False

from dotenv import load_dotenv

load_dotenv()

# Config paths
QUAD_CONFIG_DIR = Path.home() / ".quad"
QUAD_CONFIG_FILE = QUAD_CONFIG_DIR / "config.json"
QUAD_DRAFTS_DIR = QUAD_CONFIG_DIR / "drafts"

# Database config
DB_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "port": int(os.getenv("DB_PORT", 14201)),
    "dbname": os.getenv("DB_NAME", "quad_dev_db"),
    "user": os.getenv("DB_USER", "quad_user"),
    "password": os.getenv("DB_PASSWORD", "quad_dev_pass"),
}


# ─────────────────────────────────────────────────────────────
# Console Helpers
# ─────────────────────────────────────────────────────────────

class Console:
    """Simple console utilities for interactive prompts"""

    @staticmethod
    def header(text: str):
        print(f"\n  {text}")
        print(f"  {'─' * len(text)}\n")

    @staticmethod
    def success(text: str):
        print(f"  ✓ {text}")

    @staticmethod
    def info(text: str):
        print(f"  → {text}")

    @staticmethod
    def error(text: str):
        print(f"  ✗ {text}")

    @staticmethod
    def ask(question: str, default: str = None) -> str:
        """Ask a text question"""
        if default:
            prompt = f"  {question} [{default}]: "
        else:
            prompt = f"  {question}: "

        answer = input(prompt).strip()
        return answer if answer else default

    @staticmethod
    def confirm(question: str, default: bool = True) -> bool:
        """Ask yes/no question"""
        suffix = "[Y/n]" if default else "[y/N]"
        answer = input(f"  {question} {suffix}: ").strip().lower()

        if not answer:
            return default
        return answer in ('y', 'yes')

    @staticmethod
    def select(question: str, options: List[str], default: int = 0) -> int:
        """Select from numbered options"""
        print(f"  {question}")
        for i, opt in enumerate(options):
            marker = ">" if i == default else " "
            print(f"    {marker} [{i+1}] {opt}")

        while True:
            answer = input(f"  Select [1-{len(options)}]: ").strip()
            if not answer:
                return default
            try:
                idx = int(answer) - 1
                if 0 <= idx < len(options):
                    return idx
            except ValueError:
                pass
            print(f"  Please enter 1-{len(options)}")

    @staticmethod
    def multi_select(question: str, options: List[str], defaults: List[int] = None) -> List[int]:
        """Multi-select with checkboxes"""
        defaults = defaults or []
        selected = set(defaults)

        print(f"  {question}")
        for i, opt in enumerate(options):
            marker = "[x]" if i in selected else "[ ]"
            print(f"    {marker} {i+1}. {opt}")

        answer = input("  Enter numbers (comma-separated): ").strip()
        if not answer:
            return list(selected)

        try:
            indices = [int(x.strip()) - 1 for x in answer.split(',')]
            return [i for i in indices if 0 <= i < len(options)]
        except ValueError:
            return list(selected)


# ─────────────────────────────────────────────────────────────
# Excel Parser
# ─────────────────────────────────────────────────────────────

class ExcelParser:
    """Parse QUAD organization Excel file"""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.workbook = load_workbook(filepath, data_only=True)

    def get_sheet_names(self) -> List[str]:
        return self.workbook.sheetnames

    def parse_overview(self) -> Dict:
        """Parse Overview tab (Tab 1)"""
        sheet = self._get_sheet(['Overview', 'Org', 'Organization'])
        if not sheet:
            return {}

        data = {}
        for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] and row[1]:
                key = str(row[0]).lower().replace(' ', '_')
                data[key] = row[1]

        return data

    def parse_resources(self) -> List[Dict]:
        """Parse Resources tab (Tab 2)"""
        sheet = self._get_sheet(['Resources', 'Team', 'Team Members', 'People'])
        if not sheet:
            return []

        return self._parse_table(sheet)

    def parse_projects(self) -> List[Dict]:
        """Parse all Project tabs (Tab 3, 4, ...)"""
        projects = []

        for name in self.workbook.sheetnames:
            if name.lower().startswith('project') or name.lower() in ['suma', 'nutrinne', 'nutri']:
                sheet = self.workbook[name]
                project = self._parse_key_value_sheet(sheet)
                project['_sheet_name'] = name
                projects.append(project)

        return projects

    def _get_sheet(self, names: List[str]):
        """Get sheet by multiple possible names"""
        for name in names:
            if name in self.workbook.sheetnames:
                return self.workbook[name]
            # Case-insensitive search
            for sheet_name in self.workbook.sheetnames:
                if sheet_name.lower() == name.lower():
                    return self.workbook[sheet_name]
        return None

    def _parse_table(self, sheet) -> List[Dict]:
        """Parse a table with header row"""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).lower().replace(' ', '_') if h else f'col_{i}'
                   for i, h in enumerate(rows[0])]

        data = []
        for row in rows[1:]:
            if any(row):  # Skip empty rows
                item = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                data.append(item)

        return data

    def _parse_key_value_sheet(self, sheet) -> Dict:
        """Parse a key-value format sheet"""
        data = {}
        for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] and row[1]:
                key = str(row[0]).lower().replace(' ', '_')
                data[key] = row[1]
        return data


# ─────────────────────────────────────────────────────────────
# Interactive Init
# ─────────────────────────────────────────────────────────────

class QuadInit:
    """Interactive QUAD project initialization"""

    PROJECT_TYPES = ['Web Application', 'Mobile App', 'API/Backend', 'Desktop App', 'Library/Package']
    FRONTEND_TECH = ['React.js', 'Next.js', 'Vue.js', 'Angular', 'Swift (iOS)', 'Kotlin (Android)', 'None']
    BACKEND_TECH = ['Node.js', 'Python/FastAPI', 'Java/Spring Boot', 'Go', 'None']
    DATABASE_TECH = ['PostgreSQL', 'MySQL', 'MongoDB', 'SQLite', 'None']
    DELIVERABLES = ['Web Application', 'API Server', 'Mobile App', 'JAR File', 'Docker Image', 'Documentation', 'SDK/Library']

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.parser = ExcelParser(excel_path)
        self.org_data = {}
        self.resources = []
        self.projects = []

    def run(self):
        """Run the interactive initialization"""
        print("\n" + "=" * 50)
        print("  QUAD Project Initialization")
        print("=" * 50)

        # Step 1: Parse Excel
        self._parse_excel()

        # Step 2: Confirm organization
        self._confirm_organization()

        # Step 3: Confirm resources
        self._confirm_resources()

        # Step 4: Interactive project setup
        self._setup_projects()

        # Step 5: Generate plan
        plan = self._generate_plan()

        # Step 6: Save to database
        if Console.confirm("\n  Save to database?"):
            self._save_to_database()
            Console.success("Projects initialized!")
        else:
            Console.info("Skipped database save")

        # Show summary
        self._show_summary(plan)

    def _parse_excel(self):
        """Parse the Excel file"""
        Console.header("Parsing Excel File")
        Console.info(f"File: {self.excel_path}")

        sheets = self.parser.get_sheet_names()
        Console.info(f"Found {len(sheets)} sheets: {', '.join(sheets)}")

        self.org_data = self.parser.parse_overview()
        self.resources = self.parser.parse_resources()
        self.projects = self.parser.parse_projects()

        Console.success(f"Organization: {self.org_data.get('org_name', 'Unknown')}")
        Console.success(f"Resources: {len(self.resources)} team members")
        Console.success(f"Projects: {len(self.projects)} found")

    def _confirm_organization(self):
        """Confirm organization details"""
        Console.header("Organization Details")

        org_name = Console.ask("Organization name", self.org_data.get('org_name', ''))
        org_code = Console.ask("Organization code", self.org_data.get('org_code', org_name[:4].upper() if org_name else ''))
        timezone = Console.ask("Timezone", self.org_data.get('timezone', 'America/New_York'))

        self.org_data['org_name'] = org_name
        self.org_data['org_code'] = org_code
        self.org_data['timezone'] = timezone

    def _confirm_resources(self):
        """Confirm team resources"""
        Console.header("Team Resources")

        if self.resources:
            print("  Found team members:")
            for r in self.resources:
                name = r.get('name', r.get('email', 'Unknown'))
                role = r.get('role', r.get('job_title', 'Team Member'))
                skills = r.get('skills', '')
                print(f"    • {name} - {role}" + (f" ({skills})" if skills else ""))

            if not Console.confirm("\n  Is this list correct?"):
                Console.info("You can edit the Excel and re-run quad-init")
        else:
            Console.info("No team members found in Excel")
            Console.info("Add a 'Resources' or 'Team' sheet with columns: name, email, role, skills")

    def _setup_projects(self):
        """Interactive project setup"""
        Console.header("Project Setup")

        for i, project in enumerate(self.projects):
            print(f"\n  ── Project {i+1}: {project.get('project_name', project.get('_sheet_name', 'Unknown'))} ──\n")

            # Basic info
            project['name'] = Console.ask("Project name", project.get('project_name', project.get('_sheet_name', '')))
            project['description'] = Console.ask("Description", project.get('description', ''))

            # Project type
            type_idx = Console.select(
                "What type of project is this?",
                self.PROJECT_TYPES,
                self._find_index(self.PROJECT_TYPES, project.get('type', ''))
            )
            project['type'] = self.PROJECT_TYPES[type_idx]

            # Tech stack
            print("\n  Tech Stack:")

            fe_idx = Console.select(
                "Frontend technology?",
                self.FRONTEND_TECH,
                self._find_index(self.FRONTEND_TECH, project.get('frontend', ''))
            )
            project['frontend'] = self.FRONTEND_TECH[fe_idx]

            be_idx = Console.select(
                "Backend technology?",
                self.BACKEND_TECH,
                self._find_index(self.BACKEND_TECH, project.get('backend', ''))
            )
            project['backend'] = self.BACKEND_TECH[be_idx]

            db_idx = Console.select(
                "Database?",
                self.DATABASE_TECH,
                self._find_index(self.DATABASE_TECH, project.get('database', ''))
            )
            project['database'] = self.DATABASE_TECH[db_idx]

            # Deliverables
            print("\n  Deliverables:")
            del_indices = Console.multi_select(
                "What needs to be delivered?",
                self.DELIVERABLES,
                self._find_deliverable_indices(project)
            )
            project['deliverables'] = [self.DELIVERABLES[i] for i in del_indices]

            # Timeline
            print("\n  Timeline:")
            deadline_str = project.get('deadline', '')
            if deadline_str:
                project['deadline'] = Console.ask("Deadline", str(deadline_str))
            else:
                project['deadline'] = Console.ask("Deadline (YYYY-MM-DD)",
                    (datetime.now() + timedelta(days=60)).strftime('%Y-%m-%d'))

            # Owner
            if self.resources:
                owner_names = [r.get('name', r.get('email', 'Unknown')) for r in self.resources]
                owner_idx = Console.select("Project owner?", owner_names, 0)
                project['owner'] = self.resources[owner_idx]
            else:
                project['owner_email'] = Console.ask("Owner email", project.get('owner_email', ''))

            Console.success(f"Project '{project['name']}' configured")

    def _find_index(self, options: List[str], value: str) -> int:
        """Find index of value in options (case-insensitive partial match)"""
        if not value:
            return 0
        value_lower = str(value).lower()
        for i, opt in enumerate(options):
            if value_lower in opt.lower():
                return i
        return 0

    def _find_deliverable_indices(self, project: Dict) -> List[int]:
        """Find deliverable indices from project data"""
        indices = []
        deliverable = project.get('deliverable', '').lower()

        if 'jar' in deliverable:
            indices.append(3)  # JAR File
        if 'doc' in deliverable:
            indices.append(5)  # Documentation
        if 'web' in project.get('type', '').lower():
            indices.append(0)  # Web Application
        if 'api' in project.get('type', '').lower():
            indices.append(1)  # API Server

        return indices if indices else [0]

    def _generate_plan(self) -> Dict:
        """Generate project plan"""
        Console.header("Generating Plan")

        plan = {
            'org': self.org_data,
            'projects': [],
            'timeline_weeks': 8,
            'generated_at': datetime.now().isoformat()
        }

        for project in self.projects:
            # Calculate weeks until deadline
            try:
                deadline = datetime.strptime(str(project.get('deadline', '')).split()[0], '%Y-%m-%d')
                weeks = max(1, (deadline - datetime.now()).days // 7)
            except:
                weeks = 8

            # Generate phases based on project type
            phases = self._generate_phases(project, weeks)

            project_plan = {
                'name': project['name'],
                'type': project['type'],
                'deadline': project.get('deadline'),
                'deliverables': project.get('deliverables', []),
                'phases': phases
            }
            plan['projects'].append(project_plan)

        return plan

    def _generate_phases(self, project: Dict, weeks: int) -> List[Dict]:
        """Generate project phases"""
        phases = []
        deliverables = project.get('deliverables', [])

        # Standard phases based on project type
        if project['type'] == 'Web Application':
            phases = [
                {'name': 'Foundation', 'weeks': f'1-{weeks//4}', 'tasks': ['Setup project', 'Database schema', 'Auth']},
                {'name': 'Core Features', 'weeks': f'{weeks//4+1}-{weeks//2}', 'tasks': ['Main functionality', 'API endpoints']},
                {'name': 'UI/UX', 'weeks': f'{weeks//2+1}-{3*weeks//4}', 'tasks': ['Components', 'Styling', 'Responsive']},
                {'name': 'Testing & Deploy', 'weeks': f'{3*weeks//4+1}-{weeks}', 'tasks': ['Testing', 'Documentation', 'Deploy']},
            ]
        elif project['type'] == 'API/Backend':
            phases = [
                {'name': 'Setup', 'weeks': f'1-{weeks//4}', 'tasks': ['Project setup', 'Database', 'Auth']},
                {'name': 'API Development', 'weeks': f'{weeks//4+1}-{weeks//2}', 'tasks': ['Endpoints', 'Business logic']},
                {'name': 'Build Artifacts', 'weeks': f'{weeks//2+1}-{3*weeks//4}', 'tasks': ['JAR/Docker', 'CI/CD']},
                {'name': 'Documentation', 'weeks': f'{3*weeks//4+1}-{weeks}', 'tasks': ['API docs', 'Deploy guide']},
            ]
        else:
            phases = [
                {'name': 'Planning', 'weeks': f'1-{weeks//4}', 'tasks': ['Requirements', 'Design']},
                {'name': 'Development', 'weeks': f'{weeks//4+1}-{3*weeks//4}', 'tasks': ['Implementation']},
                {'name': 'Finalization', 'weeks': f'{3*weeks//4+1}-{weeks}', 'tasks': ['Testing', 'Documentation']},
            ]

        # Add deliverable-specific tasks
        if 'JAR File' in deliverables:
            phases[-2]['tasks'].append('Build JAR artifact')
        if 'Documentation' in deliverables:
            phases[-1]['tasks'].extend(['API Documentation', 'Deployment Guide'])

        return phases

    def _save_to_database(self):
        """Save projects to database"""
        try:
            conn = psycopg.connect(
                host=DB_CONFIG["host"],
                port=DB_CONFIG["port"],
                dbname=DB_CONFIG["dbname"],
                user=DB_CONFIG["user"],
                password=DB_CONFIG["password"],
                row_factory=dict_row
            )

            with conn.cursor() as cur:
                # Get or create organization
                cur.execute("""
                    SELECT id FROM quad_organizations WHERE slug = %s
                """, (self.org_data.get('org_code', '').lower(),))
                org = cur.fetchone()

                if org:
                    org_id = org['id']
                else:
                    Console.info("Organization not found, skipping DB save")
                    Console.info("Run database setup first to create org")
                    return

                # Update/insert projects (domains)
                for project in self.projects:
                    slug = project['name'].lower().replace(' ', '-')[:20]

                    cur.execute("""
                        INSERT INTO quad_domains (name, slug, description, methodology, company_id, is_active)
                        VALUES (%s, %s, %s, 'quad', %s, true)
                        ON CONFLICT (slug) DO UPDATE SET
                            name = EXCLUDED.name,
                            description = EXCLUDED.description
                        RETURNING id
                    """, (project['name'], slug, project.get('description', ''), org_id))

                    domain_id = cur.fetchone()['id']
                    Console.success(f"Saved project: {project['name']} (ID: {domain_id})")

            conn.commit()
            conn.close()

        except Exception as e:
            Console.error(f"Database error: {e}")

    def _show_summary(self, plan: Dict):
        """Show final summary"""
        Console.header("Summary for Next 2 Months")

        for project in plan['projects']:
            print(f"\n  {project['name']} ({plan['timeline_weeks']} weeks)")
            for phase in project['phases']:
                print(f"  ├── Week {phase['weeks']}: {phase['name']}")
                for task in phase['tasks'][:2]:  # Show first 2 tasks
                    print(f"  │     • {task}")

            if project['deliverables']:
                print(f"  │")
                print(f"  └── Deliverables:")
                for d in project['deliverables']:
                    print(f"        • {d}")

        print("\n" + "=" * 50)


# ─────────────────────────────────────────────────────────────
# Config Management
# ─────────────────────────────────────────────────────────────

def load_config() -> Optional[Dict]:
    """Load existing QUAD config"""
    if QUAD_CONFIG_FILE.exists():
        try:
            return json.loads(QUAD_CONFIG_FILE.read_text())
        except:
            return None
    return None


def save_config(config: Dict):
    """Save QUAD config"""
    QUAD_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    QUAD_CONFIG_FILE.write_text(json.dumps(config, indent=2))


def load_draft(name: str) -> Optional[Dict]:
    """Load a saved draft"""
    draft_file = QUAD_DRAFTS_DIR / f"{name}.json"
    if draft_file.exists():
        try:
            return json.loads(draft_file.read_text())
        except:
            return None
    return None


def save_draft(name: str, data: Dict):
    """Save work as draft"""
    QUAD_DRAFTS_DIR.mkdir(parents=True, exist_ok=True)
    draft_file = QUAD_DRAFTS_DIR / f"{name}.json"
    data['_saved_at'] = datetime.now().isoformat()
    draft_file.write_text(json.dumps(data, indent=2))
    Console.success(f"Draft saved: {draft_file}")


def list_drafts() -> List[str]:
    """List available drafts"""
    if not QUAD_DRAFTS_DIR.exists():
        return []
    return [f.stem for f in QUAD_DRAFTS_DIR.glob("*.json")]


# ─────────────────────────────────────────────────────────────
# Pure Interactive Mode (No Excel)
# ─────────────────────────────────────────────────────────────

class QuadInteractiveInit:
    """Pure interactive initialization without Excel"""

    PROJECT_TYPES = ['Web Application', 'Mobile App', 'API/Backend', 'Desktop App', 'Library/Package']
    FRONTEND_TECH = ['React.js', 'Next.js', 'Vue.js', 'Angular', 'Swift (iOS)', 'Kotlin (Android)', 'None']
    BACKEND_TECH = ['Node.js', 'Python/FastAPI', 'Java/Spring Boot', 'Go', 'None']
    DATABASE_TECH = ['PostgreSQL', 'MySQL', 'MongoDB', 'SQLite', 'None']
    DELIVERABLES = ['Web Application', 'API Server', 'Mobile App', 'JAR File', 'Docker Image', 'Documentation', 'SDK/Library']

    def __init__(self, resume_draft: str = None):
        self.resume_draft = resume_draft
        self.org_data = {}
        self.resources = []
        self.projects = []

        # Load existing config if available
        self.config = load_config()

        # Load draft if resuming
        if resume_draft:
            draft = load_draft(resume_draft)
            if draft:
                self.org_data = draft.get('org', {})
                self.resources = draft.get('resources', [])
                self.projects = draft.get('projects', [])
                Console.success(f"Resumed draft: {resume_draft}")

    def run(self):
        """Run pure interactive initialization"""
        print("\n" + "=" * 50)
        print("  QUAD Interactive Project Setup")
        print("=" * 50)

        # Check for existing config
        if self.config and not self.resume_draft:
            self._show_existing_config()

        # Check for available drafts
        if not self.resume_draft:
            drafts = list_drafts()
            if drafts:
                Console.header("Found Saved Drafts")
                for i, d in enumerate(drafts):
                    print(f"    [{i+1}] {d}")
                if Console.confirm("\n  Resume a draft?", default=False):
                    idx = Console.select("Which draft?", drafts)
                    draft = load_draft(drafts[idx])
                    if draft:
                        self.org_data = draft.get('org', {})
                        self.resources = draft.get('resources', [])
                        self.projects = draft.get('projects', [])
                        Console.success(f"Loaded draft: {drafts[idx]}")

        # Step 1: Organization setup
        self._setup_organization()

        # Step 2: Team setup (optional)
        if Console.confirm("\n  Add team members?", default=False):
            self._setup_resources()

        # Step 3: Project setup
        self._setup_projects_interactive()

        # Save draft option
        if Console.confirm("\n  Save as draft? (can resume later)"):
            draft_name = Console.ask("Draft name", self.projects[0]['name'].lower().replace(' ', '-') if self.projects else 'project')
            save_draft(draft_name, {
                'org': self.org_data,
                'resources': self.resources,
                'projects': self.projects
            })

        # Step 4: Generate or suggest Excel
        self._finalize()

    def _show_existing_config(self):
        """Show existing QUAD config"""
        Console.header("Existing Configuration Found")
        print(f"  Organization: {self.config.get('org_name', 'Unknown')}")
        print(f"  Config file: {QUAD_CONFIG_FILE}")

        if Console.confirm("\n  Use existing organization?"):
            self.org_data = self.config
        else:
            Console.info("Starting fresh setup")

    def _setup_organization(self):
        """Setup organization interactively"""
        Console.header("Organization Setup")

        self.org_data['org_name'] = Console.ask(
            "Organization name",
            self.org_data.get('org_name', 'My Company')
        )
        self.org_data['org_code'] = Console.ask(
            "Organization code (short)",
            self.org_data.get('org_code', self.org_data['org_name'][:4].upper())
        )
        self.org_data['timezone'] = Console.ask(
            "Timezone",
            self.org_data.get('timezone', 'America/New_York')
        )

        # Save to config
        if Console.confirm("  Save as default organization?"):
            save_config(self.org_data)
            Console.success("Saved to ~/.quad/config.json")

    def _setup_resources(self):
        """Setup team members interactively"""
        Console.header("Team Members")

        while True:
            print("\n  Add team member:")
            name = Console.ask("Name")
            if not name:
                break

            email = Console.ask("Email")
            role = Console.ask("Role", "Developer")
            skills = Console.ask("Skills (comma-separated)", "")

            self.resources.append({
                'name': name,
                'email': email,
                'role': role,
                'skills': skills
            })
            Console.success(f"Added: {name}")

            if not Console.confirm("  Add another?", default=False):
                break

    def _setup_projects_interactive(self):
        """Setup projects purely interactively"""
        Console.header("Project Setup")

        while True:
            project = {}

            print("\n  ── New Project ──\n")

            # Basic info
            project['name'] = Console.ask("Project name")
            if not project['name']:
                if not self.projects:
                    Console.error("At least one project required")
                    continue
                break

            project['description'] = Console.ask("Description", "")

            # Project type
            type_idx = Console.select("What type of project?", self.PROJECT_TYPES)
            project['type'] = self.PROJECT_TYPES[type_idx]

            # Tech stack
            print("\n  Tech Stack:")

            fe_idx = Console.select("Frontend technology?", self.FRONTEND_TECH)
            project['frontend'] = self.FRONTEND_TECH[fe_idx]

            be_idx = Console.select("Backend technology?", self.BACKEND_TECH)
            project['backend'] = self.BACKEND_TECH[be_idx]

            db_idx = Console.select("Database?", self.DATABASE_TECH)
            project['database'] = self.DATABASE_TECH[db_idx]

            # Deliverables
            print("\n  Deliverables:")
            del_indices = Console.multi_select("What needs to be delivered?", self.DELIVERABLES, [0])
            project['deliverables'] = [self.DELIVERABLES[i] for i in del_indices]

            # Timeline
            project['deadline'] = Console.ask(
                "Deadline (YYYY-MM-DD)",
                (datetime.now() + timedelta(days=60)).strftime('%Y-%m-%d')
            )

            # Owner
            if self.resources:
                owner_names = [r.get('name', 'Unknown') for r in self.resources]
                owner_idx = Console.select("Project owner?", owner_names)
                project['owner'] = self.resources[owner_idx]
            else:
                project['owner_email'] = Console.ask("Owner email", "")

            self.projects.append(project)
            Console.success(f"Project '{project['name']}' added")

            if not Console.confirm("\n  Add another project?", default=False):
                break

    def _finalize(self):
        """Finalize - suggest Excel or save to DB"""
        Console.header("Finalization")

        # Option 1: Generate Excel template with filled data
        if HAS_OPENPYXL:
            if Console.confirm("  Generate Excel file with your data?"):
                self._generate_filled_excel()
                return

        # Option 2: Save directly to database
        if HAS_PSYCOPG:
            if Console.confirm("  Save directly to database?"):
                self._save_to_database()
                Console.success("Projects initialized!")
                return

        # Option 3: Just show summary
        Console.info("Your project configuration:")
        self._show_summary()

        Console.info("\nNext steps:")
        Console.info("  1. Run 'quad template' to create Excel file")
        Console.info("  2. Fill in Excel with your project details")
        Console.info("  3. Run 'quad-init @your-file.xlsx' to initialize")

    def _generate_filled_excel(self):
        """Generate Excel pre-filled with interactive data"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # Overview sheet
        ws = wb.active
        ws.title = "Overview"
        overview_data = [
            ("Field", "Value"),
            ("org_name", self.org_data.get('org_name', '')),
            ("org_code", self.org_data.get('org_code', '')),
            ("timezone", self.org_data.get('timezone', '')),
        ]
        for row in overview_data:
            ws.append(row)

        # Resources sheet
        ws = wb.create_sheet("Resources")
        ws.append(["Name", "Email", "Role", "Skills"])
        for r in self.resources:
            ws.append([r.get('name'), r.get('email'), r.get('role'), r.get('skills')])

        # Project sheets
        for i, project in enumerate(self.projects):
            ws = wb.create_sheet(f"Project {i+1}")
            project_data = [
                ("project_name", project.get('name', '')),
                ("description", project.get('description', '')),
                ("type", project.get('type', '')),
                ("frontend", project.get('frontend', '')),
                ("backend", project.get('backend', '')),
                ("database", project.get('database', '')),
                ("deliverables", ', '.join(project.get('deliverables', []))),
                ("deadline", project.get('deadline', '')),
                ("owner_email", project.get('owner_email', '')),
            ]
            for row in project_data:
                ws.append(row)

        # Save
        filename = f"quad-{self.org_data.get('org_code', 'org').lower()}-setup.xlsx"
        wb.save(filename)
        Console.success(f"Excel file created: {filename}")
        Console.info("You can edit this file and re-run 'quad-init @{filename}'")

    def _save_to_database(self):
        """Save to database"""
        if not HAS_PSYCOPG:
            Console.error("psycopg not installed. Run: pip install psycopg[binary]")
            return

        try:
            conn = psycopg.connect(
                host=DB_CONFIG["host"],
                port=DB_CONFIG["port"],
                dbname=DB_CONFIG["dbname"],
                user=DB_CONFIG["user"],
                password=DB_CONFIG["password"],
                row_factory=dict_row
            )

            with conn.cursor() as cur:
                # Get or create organization
                cur.execute("""
                    SELECT id FROM quad_organizations WHERE slug = %s
                """, (self.org_data.get('org_code', '').lower(),))
                org = cur.fetchone()

                if org:
                    org_id = org['id']
                else:
                    Console.info("Organization not found in database")
                    Console.info("Run database setup first")
                    return

                # Save projects
                for project in self.projects:
                    slug = project['name'].lower().replace(' ', '-')[:20]

                    cur.execute("""
                        INSERT INTO quad_domains (name, slug, description, methodology, company_id, is_active)
                        VALUES (%s, %s, %s, 'quad', %s, true)
                        ON CONFLICT (slug) DO UPDATE SET
                            name = EXCLUDED.name,
                            description = EXCLUDED.description
                        RETURNING id
                    """, (project['name'], slug, project.get('description', ''), org_id))

                    domain_id = cur.fetchone()['id']
                    Console.success(f"Saved: {project['name']} (ID: {domain_id})")

            conn.commit()
            conn.close()

        except Exception as e:
            Console.error(f"Database error: {e}")

    def _show_summary(self):
        """Show configuration summary"""
        print(f"\n  Organization: {self.org_data.get('org_name')}")
        print(f"  Code: {self.org_data.get('org_code')}")
        print(f"  Team: {len(self.resources)} members")
        print(f"  Projects: {len(self.projects)}")

        for p in self.projects:
            print(f"\n  ─ {p['name']} ({p['type']})")
            print(f"    Frontend: {p.get('frontend', 'N/A')}")
            print(f"    Backend: {p.get('backend', 'N/A')}")
            print(f"    Database: {p.get('database', 'N/A')}")
            print(f"    Deadline: {p.get('deadline', 'TBD')}")


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────

def show_help():
    """Show usage help"""
    print("""
QUAD Project Initialization
============================

Usage:
  quad-init                      Interactive mode (no file needed)
  quad-init @<excel-file>        Initialize from Excel file
  quad-init --resume <name>      Resume a saved draft
  quad-init --list-drafts        List saved drafts

Examples:
  quad-init                      Start interactive setup
  quad-init @org-setup.xlsx      Load from Excel
  quad-init --resume bank-demo   Resume 'bank-demo' draft

Interactive mode will:
  1. Check for existing ~/.quad/config.json
  2. Check for saved drafts in ~/.quad/drafts/
  3. Walk you through project setup
  4. Optionally generate Excel or save to database
""")


def main():
    # No arguments = interactive mode
    if len(sys.argv) < 2:
        init = QuadInteractiveInit()
        init.run()
        return

    # Parse arguments
    arg = sys.argv[1]

    # Help
    if arg in ('-h', '--help', 'help'):
        show_help()
        return

    # List drafts
    if arg == '--list-drafts':
        drafts = list_drafts()
        if drafts:
            Console.header("Saved Drafts")
            for d in drafts:
                print(f"    • {d}")
        else:
            Console.info("No drafts found")
        return

    # Resume draft
    if arg == '--resume' and len(sys.argv) > 2:
        draft_name = sys.argv[2]
        init = QuadInteractiveInit(resume_draft=draft_name)
        init.run()
        return

    # Excel file mode
    if arg.startswith('@'):
        filepath = arg[1:]
    elif arg in ('--file', '-f') and len(sys.argv) > 2:
        filepath = sys.argv[2]
    else:
        filepath = arg

    # Check openpyxl for Excel mode
    if not HAS_OPENPYXL:
        Console.error("openpyxl not installed for Excel mode")
        Console.info("Run: pip install openpyxl")
        Console.info("Or use: quad-init (interactive mode)")
        sys.exit(1)

    # Check file exists
    if not os.path.exists(filepath):
        Console.error(f"File not found: {filepath}")
        sys.exit(1)

    # Run Excel-based initialization
    init = QuadInit(filepath)
    init.run()


if __name__ == "__main__":
    main()
